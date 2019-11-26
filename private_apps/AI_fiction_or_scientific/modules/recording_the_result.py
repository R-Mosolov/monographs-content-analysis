from openpyxl import Workbook
from openpyxl.styles import Font

from modules import sentences_number_in_text
from modules import words_number_in_text
from modules import punctuation_signs_number_in_text
from modules import relative_punctuation_signs_number_in_text
from modules import words_number_mean_in_sentence
from modules import letters_number_mean_in_word


def record():
    wb = Workbook()
    ws = wb.active

    # Style

    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)

    # Structure

    ws['A1'] = 'Индикатор'
    ws['B1'] = 'Худ. книги'
    ws['C1'] = 'Науч. книги'

    ws['A2'] = 'Количество предложений в тексте (ед.)'
    ws['B2'] = sentences_number_in_text.get()
    ws['C2'] = 'Cell is empty'

    ws['A3'] = 'Количество слов в тексте (ед.)'
    ws['B3'] = words_number_in_text.get()
    ws['C3'] = 'Cell is empty'

    ws['A4'] = 'Абсолютное количество знаков препинаний в тексте (ед.)'
    ws['B4'] = punctuation_signs_number_in_text.get()
    ws['C4'] = 'Cell is empty'

    ws['A5'] = 'Относительное количество знаков препинаний в тексте (%)'
    ws['B5'] = relative_punctuation_signs_number_in_text.get()
    ws['C5'] = 'Cell is empty'

    ws['A6'] = 'Среднее число слов в предложении (ед.)'
    ws['B6'] = words_number_mean_in_sentence.get()
    ws['C6'] = 'Cell is empty'

    ws['A7'] = 'Среднее число букв в слове (ед.)'
    ws['B7'] = letters_number_mean_in_word.get()
    ws['C7'] = 'Cell is empty'

    wb.save("result_fiction.xlsx")
    print('Результат анализа успешно записан в файл!')