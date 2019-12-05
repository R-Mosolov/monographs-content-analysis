from openpyxl import Workbook
from openpyxl.styles import Font

from modules.additional_indicators import text__sentence
from modules.additional_indicators import text__word
from modules.additional_indicators import text__sign_abs

from modules.main_indicators import text__sign_rel
from modules.main_indicators import sentence__word
from modules.main_indicators import word__letter


def run():
    wb = Workbook()
    ws = wb.active

    # Style

    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)

    # Structure

    ws['A1'] = 'Индикатор'
    ws['B1'] = 'Худ. книги'
    ws['C1'] = 'Науч. книги'

    ws['A2'] = 'Количество предложений в тексте (ед.)'
    ws['B2'] = text__sentence.calc()
    ws['C2'] = 'Cell is empty'

    ws['A3'] = 'Количество слов в тексте (ед.)'
    ws['B3'] = text__word.calc()
    ws['C3'] = 'Cell is empty'

    ws['A4'] = 'Абсолютное количество знаков препинаний в тексте (ед.)'
    ws['B4'] = text__sign_abs.calc()
    ws['C4'] = 'Cell is empty'

    ws['A5'] = 'Относительное количество знаков препинаний в тексте (%)'
    ws['B5'] = text__sign_rel.calc()
    ws['C5'] = 'Cell is empty'

    ws['A6'] = 'Среднее число слов в предложении (ед.)'
    ws['B6'] = sentence__word.calc()
    ws['C6'] = 'Cell is empty'

    ws['A7'] = 'Среднее число букв в слове (ед.)'
    ws['B7'] = word__letter.calc()
    ws['C7'] = 'Cell is empty'

    wb.save("result_new.xlsx")
    print('Результат анализа успешно записан в файл!')